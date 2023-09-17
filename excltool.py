import os
import openpyxl
from openpyxl import Workbook
import numpy as np

wb = openpyxl.load_workbook('DATA RUMAH.xlsx')
ws = wb.active

y_train_data = np.array([])
x_train_data = np.array([])

def get_x_train():
    x_train = np.array([])  # Initialize x_train as a local variable
    for x in range(2, 1003):
        data = ws.cell(row=x, column=5).value
        x_train = np.append(x_train, data)
    return x_train

def get_y_train():
    y_train = np.array([])  # Initialize y_train as a local variable
    for y in range(2, 1003):
        data = ws.cell(row=y, column=3).value
        y_train = np.append(y_train, data)
    return y_train
    
print (type(get_y_train()))



    

